import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Side, Font
from openpyxl.utils import get_column_letter
from os.path import exists
import subprocess

xlsx_name = ""

def GetUUID():
    cmd = 'wmic csproduct get uuid'
    uuid = str(subprocess.check_output(cmd))
    pos1 = uuid.find("\\n")+2
    uuid = uuid[pos1:-15]
    return uuid


def create_excel(rsrp, arsrp, rsrq, arsrq, rssi, arssi, sinr, asinr, rfm, ping, dl, up):
    row = 3
    xlsx_name = "Airbox_" + str(GetUUID()) + ".xlsx"

    if exists(xlsx_name):
        wb = openpyxl.load_workbook(xlsx_name)
        ws = wb.active
        row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Speedtest"
        try:
            wb.save(xlsx_name)
        except PermissionError:
            print("Fichier Excel déjà ouvert. Il faut le fermer.")
            input()
            exit()
    generate_template(xlsx_name, rsrp, arsrp, rsrq, arsrq, rssi, arssi, sinr, asinr, rfm, ping, dl, up, row)


def generate_template(xlsx, rsrp, arsrp, rsrq, arsrq, rssi, arssi, sinr, asinr, rfm, ping, dl, up, row):
    wb = load_workbook(xlsx)
    ws = wb.active

    # HEADER SIGNAUX
    ws.merge_cells('A1:E1')
    set_border(ws, 'A1:E1')
    set_color(ws, 'A1:E1')
    ws['A1'] = 'SIGNAUX'
    fontstyle = Font(size="14", bold=True)
    ws['A1'].font = fontstyle
    ws['A1'].alignment = Alignment(horizontal='center')

    # HEADER SPEEDTEST
    ws.merge_cells('F1:H1')
    set_border(ws, 'F1:H1')
    set_color(ws, 'F1:H1')
    ws['F1'] = 'SPEEDTEST'
    fontstyle = Font(size="14", bold=True)
    ws['F1'].font = fontstyle
    ws['F1'].alignment = Alignment(horizontal='center')

    # HEADER #2 SIGNAUX
    ws['A2'] = 'RSRP (dBm)'
    set_border(ws, 'A2:A2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'] = 'RSRP (avg)'
    set_border(ws, 'B2:B2')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'] = 'RSRQ (dBm)'
    set_border(ws, 'C2:C2')
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'] = 'RSRQ (avg)'
    set_border(ws, 'D2:D2')
    ws['D2'].alignment = Alignment(horizontal='center')
    ws['E2'] = 'RSSI (dBm)'
    set_border(ws, 'E2:E2')
    ws['E2'].alignment = Alignment(horizontal='center')
    ws['F2'] = 'RSSI (avg)'
    set_border(ws, 'F2:F2')
    ws['F2'].alignment = Alignment(horizontal='center')
    ws['G2'] = 'SINR (dB)'
    set_border(ws, 'G2:G2')
    ws['G2'].alignment = Alignment(horizontal='center')
    ws['H2'] = 'SINR (avg)'
    set_border(ws, 'H2:H2')
    ws['H2'].alignment = Alignment(horizontal='center')
    ws['I2'] = 'RF-MARGIN (%)'
    set_border(ws, 'I2:I2')
    ws['I2'].alignment = Alignment(horizontal='center')
    ws['J2'] = 'PING (ms)'
    set_border(ws, 'J2:J2')
    ws['J2'].alignment = Alignment(horizontal='center')
    ws['K2'] = 'DOWNLOAD'
    set_border(ws, 'K2:K2')
    ws['K2'].alignment = Alignment(horizontal='center')
    ws['L2'] = 'UPLOAD'
    set_border(ws, 'L2:L2')
    ws['L2'].alignment = Alignment(horizontal='center')
    ws['M2'] = 'DATE'
    set_border(ws, 'M2:M2')
    ws['M2'].alignment = Alignment(horizontal='center')

    fill_with_datas(ws, rsrp, arsrp, rsrq, arsrq, rssi, arssi, sinr, asinr, rfm, ping, dl, up, row)
    best_fit_column(ws)
    set_alignment(ws)
    wb.save(xlsx)


def fill_with_datas(ws, rsrp, arsrp, rsrq, arsrq, rssi, arssi, sinr, asinr, rfm, ping, dl, up, row):
    date = datetime.datetime.today().strftime("%d-%m-%Y_%H:%M:%S")
    ws.cell(row=row, column=1).value = rsrp
    ws.cell(row=row, column=2).value = arsrp
    ws.cell(row=row, column=3).value = rsrq
    ws.cell(row=row, column=4).value = arsrq
    ws.cell(row=row, column=5).value = rssi
    ws.cell(row=row, column=6).value = arssi
    ws.cell(row=row, column=7).value = sinr
    ws.cell(row=row, column=8).value = asinr
    ws.cell(row=row, column=9).value = rfm
    ws.cell(row=row, column=10).value = ping
    ws.cell(row=row, column=11).value = dl
    ws.cell(row=row, column=12).value = up
    ws.cell(row=row, column=13).value = date


def best_fit_column(ws):
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        ws.column_dimensions[new_column_letter].width = new_column_length * 1.23


def set_alignment(ws):
    for row in ws[1:ws.max_row]:
        for col in range(0, 13):
            cell = row[col]
            cell.alignment = Alignment(horizontal='center')


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_color(ws, cell_range):
    color = PatternFill(start_color='0070c0',
                        end_color='0070c0',
                        fill_type='solid')
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = color
